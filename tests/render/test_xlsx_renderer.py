"""Tests for render/xlsx_renderer.py."""

from __future__ import annotations

import io
import os

from openpyxl import load_workbook

from rfp_intake.domain.registry import Registry
from rfp_intake.domain.schemas import (
    Contradiction,
    FieldRecord,
    Provenance,
    ResolvedField,
    RunState,
)
from rfp_intake.render.xlsx_renderer import build_report_xlsx


def _use_real_registry(fields_yaml_path) -> None:  # type: ignore[no-untyped-def]
    os.environ["RFP_INTAKE_FIELDS_YAML_PATH"] = str(fields_yaml_path)
    from rfp_intake.domain.registry import get_registry
    get_registry.cache_clear()


def _registry() -> Registry:
    from rfp_intake.domain.registry import get_registry
    return get_registry()


class TestBuildReportXlsx:
    def test_produces_two_sheets(self, fields_yaml_path) -> None:  # type: ignore[no-untyped-def]
        _use_real_registry(fields_yaml_path)
        xlsx = build_report_xlsx(RunState(run_id="r-1"), _registry())
        wb = load_workbook(io.BytesIO(xlsx))
        assert wb.sheetnames == ["Fields", "Review Queue"]

    def test_header_row(self, fields_yaml_path) -> None:  # type: ignore[no-untyped-def]
        _use_real_registry(fields_yaml_path)
        xlsx = build_report_xlsx(RunState(run_id="r-1"), _registry())
        wb = load_workbook(io.BytesIO(xlsx))
        header = [c.value for c in wb["Fields"][1]]
        assert header == [
            "Group", "Variable", "Value", "Status", "Confidence",
            "Source Doc", "Page", "Quote", "Contradiction",
        ]

    def test_confirmed_field_row_content_and_fill(self, fields_yaml_path) -> None:  # type: ignore[no-untyped-def]
        _use_real_registry(fields_yaml_path)
        rf = ResolvedField(
            field_id="ops.sites_total", value=75, status="confirmed", confidence=0.92,
            sources=[Provenance(doc_id="rfp1", doc_kind="rfp", page=5)], quote="75 sites",
        )
        xlsx = build_report_xlsx(RunState(run_id="r-1", resolved=[rf]), _registry())
        wb = load_workbook(io.BytesIO(xlsx))
        row = [c.value for c in wb["Fields"][2]]

        assert row[0] == "operational_metrics"
        assert row[1] == "Total number of sites"
        assert row[2] == 75
        assert row[3] == "confirmed"
        assert row[4] == 0.92
        assert row[5] == "rfp1"
        assert row[6] == 5
        assert row[7] == "75 sites"

        fill = wb["Fields"][2][0].fill
        assert fill.fgColor.rgb == "00C8E6C9"  # green

    def test_conflict_row_gets_red_fill(self, fields_yaml_path) -> None:  # type: ignore[no-untyped-def]
        _use_real_registry(fields_yaml_path)
        record = FieldRecord(
            field_id="ops.sites_total", group="operational_metrics", raw_value="75",
            quote="q", provenance=Provenance(doc_id="d1", doc_kind="rfp", page=1), confidence=0.9,
        )
        contradiction = Contradiction(
            field_id="ops.sites_total", records=[record], verdict="conflict",
            explanation="mismatch", severity="high",
        )
        rf = ResolvedField(
            field_id="ops.sites_total", value=None, status="needs_review", confidence=0.5,
            contradiction=contradiction,
        )
        xlsx = build_report_xlsx(RunState(run_id="r-1", resolved=[rf]), _registry())
        wb = load_workbook(io.BytesIO(xlsx))

        fill = wb["Fields"][2][0].fill
        assert fill.fgColor.rgb == "00FFCDD2"  # red
        assert "conflict" in wb["Fields"][2][8].value

    def test_not_found_row_gets_grey_fill(self, fields_yaml_path) -> None:  # type: ignore[no-untyped-def]
        _use_real_registry(fields_yaml_path)
        rf = ResolvedField(
            field_id="ops.sites_total", value=None, status="not_found", confidence=0.0,
        )
        xlsx = build_report_xlsx(RunState(run_id="r-1", resolved=[rf]), _registry())
        wb = load_workbook(io.BytesIO(xlsx))
        fill = wb["Fields"][2][0].fill
        assert fill.fgColor.rgb == "00E0E0E0"  # grey

    def test_review_queue_excludes_confirmed(self, fields_yaml_path) -> None:  # type: ignore[no-untyped-def]
        _use_real_registry(fields_yaml_path)
        confirmed = ResolvedField(
            field_id="ops.sites_total", value=75, status="confirmed", confidence=0.95,
        )
        needs_review = ResolvedField(
            field_id="study.indication", value="oncology", status="needs_review", confidence=0.6,
        )
        xlsx = build_report_xlsx(
            RunState(run_id="r-1", resolved=[confirmed, needs_review]), _registry(),
        )
        wb = load_workbook(io.BytesIO(xlsx))

        review_field_ids = [row[1].value for row in wb["Review Queue"].iter_rows(min_row=2)]
        assert "Total number of sites" not in review_field_ids
        assert "Disease type / indication" in review_field_ids

    def test_review_queue_puts_budget_drivers_first(self, fields_yaml_path) -> None:  # type: ignore[no-untyped-def]
        _use_real_registry(fields_yaml_path)
        # study.indication: not a budget driver. ops.sites_total: budget_driver: true.
        non_driver = ResolvedField(
            field_id="study.indication", value="x", status="needs_review", confidence=0.5,
        )
        driver = ResolvedField(
            field_id="ops.sites_total", value=75, status="needs_review", confidence=0.5,
        )
        xlsx = build_report_xlsx(
            RunState(run_id="r-1", resolved=[non_driver, driver]), _registry(),
        )
        wb = load_workbook(io.BytesIO(xlsx))

        first_data_row = [c.value for c in wb["Review Queue"][2]]
        assert first_data_row[1] == "Total number of sites"

    def test_empty_state_still_has_headers(self, fields_yaml_path) -> None:  # type: ignore[no-untyped-def]
        _use_real_registry(fields_yaml_path)
        xlsx = build_report_xlsx(RunState(run_id="r-1"), _registry())
        wb = load_workbook(io.BytesIO(xlsx))
        assert wb["Fields"].max_row == 1
        assert wb["Review Queue"].max_row == 1
