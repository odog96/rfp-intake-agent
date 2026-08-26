"""report.xlsx — spreadsheet renderer alongside the primary JSON/PDF outputs.
Pure function returning bytes; no filesystem access. See ARCHITECTURE.md §4.10.

One row per field: Group | Variable | Value | Status | Confidence |
Source Doc | Page | Quote | Contradiction, conditional-formatted
red/amber/green/grey, plus a Review Queue sheet (non-confirmed rows,
budget drivers first).
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from rfp_intake.domain.registry import FieldDef, Registry
from rfp_intake.domain.schemas import ResolvedField, RunState

_HEADERS = [
    "Group", "Variable", "Value", "Status", "Confidence",
    "Source Doc", "Page", "Quote", "Contradiction",
]

_FILL_CONFIRMED = PatternFill("solid", fgColor="C8E6C9")   # green
_FILL_NEEDS_REVIEW = PatternFill("solid", fgColor="FFE0B2")  # amber
_FILL_CONFLICT = PatternFill("solid", fgColor="FFCDD2")   # red
_FILL_NOT_FOUND = PatternFill("solid", fgColor="E0E0E0")   # grey

_COLUMN_WIDTHS = [20, 32, 24, 14, 12, 20, 6, 50, 40]


def _row_fill(rf: ResolvedField) -> PatternFill:
    if rf.contradiction is not None and rf.contradiction.verdict == "conflict":
        return _FILL_CONFLICT
    if rf.status == "confirmed":
        return _FILL_CONFIRMED
    if rf.status == "needs_review":
        return _FILL_NEEDS_REVIEW
    return _FILL_NOT_FOUND  # not_found, not_specified


def _row_values(rf: ResolvedField, field_def: FieldDef | None) -> list[Any]:
    label = field_def.label if field_def else rf.field_id
    group = field_def.group if field_def else ""
    source = rf.sources[0] if rf.sources else None
    contradiction_text = ""
    if rf.contradiction is not None:
        verdict = rf.contradiction.verdict or "pending"
        explanation = rf.contradiction.explanation or ""
        contradiction_text = f"{verdict}: {explanation}"

    return [
        group,
        f"{label} [{rf.scope}]" if rf.scope else label,
        rf.value if rf.value is not None else "",
        rf.status,
        round(rf.confidence, 2),
        source.doc_id if source else "",
        source.page if source else "",
        rf.quote or "",
        contradiction_text,
    ]


def _write_sheet(ws: Worksheet, rows: list[tuple[ResolvedField, FieldDef | None]]) -> None:
    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for rf, field_def in rows:
        ws.append(_row_values(rf, field_def))
        fill = _row_fill(rf)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    for i, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    ws.freeze_panes = "A2"


def build_report_xlsx(state: RunState, registry: Registry) -> bytes:
    """Render the XLSX report as an in-memory workbook, returned as bytes."""
    field_by_id = {f.id: f for f in registry.fields}
    budget_drivers = {f.id for f in registry.fields if f.budget_driver}

    all_rows = [(rf, field_by_id.get(rf.field_id)) for rf in state.resolved]

    wb = Workbook()
    fields_ws = wb.active
    fields_ws.title = "Fields"
    _write_sheet(fields_ws, all_rows)

    review_rows = sorted(
        (row for row in all_rows if row[0].status != "confirmed"),
        key=lambda row: row[0].field_id not in budget_drivers,  # False (budget driver) sorts first
    )
    review_ws = wb.create_sheet("Review Queue")
    _write_sheet(review_ws, review_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
